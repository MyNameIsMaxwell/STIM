import time

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter


def space_out():
    file_name = 'Непогода по регионам.xls'
    new_file_name = f'{file_name[:-4]}(new).xlsx'

    df = pd.read_excel(file_name)

    # Заполните пустые ячейки в первом столбце данными из предыдущей строки
    df.iloc[:, 0].fillna(method='ffill', inplace=True)
    # df.iloc[:, 1].fillna(method='ffill', inplace=True)
    # df.iloc[:, 2].fillna(method='ffill', inplace=True)
    # df.iloc[:, 3].fillna(method='ffill', inplace=True)

    df.to_excel(new_file_name, index=False)

    time.sleep(1)

    workbook = load_workbook(new_file_name)
    worksheet = workbook.active

    for i in range(1, worksheet.max_column + 1):
        letter = get_column_letter(i)

        worksheet.column_dimensions[letter].width = 30

    workbook.save(new_file_name)


space_out()
