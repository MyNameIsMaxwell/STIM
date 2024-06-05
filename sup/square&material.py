import time

import gspread
import openpyxl
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import numpy as np
from pprint import pprint

from openpyxl.comments import Comment
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

scope = ['https://www.googleapis.com/auth/spreadsheets',
         "https://www.googleapis.com/auth/drive"]

credentials = ServiceAccountCredentials.from_json_keyfile_name("stim-downtime-credentials.json", scope)
client = gspread.authorize(credentials)

spreadsheet = client.open_by_key("153loT8FiuqTjECvGdFFpgyjcu5GuctrCrgO0RTELi5U")

all_comments = {}


def google_table_read():
    sheet = spreadsheet.worksheet('РБ')
    columns_to_keep = ['Контракт', 'Вид работ', 'Норма расхода(контракт)', 'Норма расхода(план)']
    data = sheet.get_all_values()
    df = pd.DataFrame(data[1:], columns=data[0])
    df = df[columns_to_keep]
    df.replace('', np.nan, inplace=True)
    df['Контракт'].fillna(method='ffill', inplace=True)
    for region in spreadsheet.worksheets()[1:]:
        sheet = spreadsheet.worksheet(region.title)
        data = sheet.get_all_values()
        df_next = pd.DataFrame(data[1:], columns=data[0])
        df_next = df_next[columns_to_keep]
        df_next.replace('', np.nan, inplace=True)
        df_next['Контракт'].fillna(method='ffill', inplace=True)
        df = pd.concat([df, df_next])

    return df


def material_inaccuracy_calc(worksheet, material_sheet_index, google_values):
    good = PatternFill(start_color="7FFF00", end_color="7FFF00", fill_type="solid")
    not_good = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    bad = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    material_values = worksheet[material_sheet_index + 1]
    if not google_values[3] in [np.nan]:
        for index, value in enumerate(material_values[:35]):
            if isinstance(value.value, float) or isinstance(value.value, int):
                inaccuracy_percent = 100 - ((value.value * 100) / float(google_values[3].replace(',', '.')))
                if inaccuracy_percent > 10 or inaccuracy_percent < -10:
                    worksheet.cell(row=material_sheet_index + 1, column=index + 1).fill = bad
                elif inaccuracy_percent > 5 or inaccuracy_percent < -5:
                    worksheet.cell(row=material_sheet_index + 1, column=index + 1).fill = not_good
                elif inaccuracy_percent <= 5 or inaccuracy_percent >= -5:
                    worksheet.cell(row=material_sheet_index + 1, column=index + 1).fill = good
    # elif not google_values[2] in [np.nan]:
    #     for index, value in enumerate(material_values[:35]):
    #         if isinstance(value.value, float) or isinstance(value.value, int):
    #             inaccuracy_percent = 100 - ((value.value * 100) / float(google_values[2].replace(',', '.')))
    #             if inaccuracy_percent > 10 or inaccuracy_percent < -10:
    #                 worksheet.cell(row=material_sheet_index + 1, column=index + 1).fill = bad
    #             elif inaccuracy_percent > 5 or inaccuracy_percent < -5:
    #                 worksheet.cell(row=material_sheet_index + 1, column=index + 1).fill = not_good
    #             elif inaccuracy_percent <= 5 or inaccuracy_percent >= -5:
    #                 worksheet.cell(row=material_sheet_index + 1, column=index + 1).fill = good
    #             comment = Comment("Примечание","")
    #             worksheet.cell(row=material_sheet_index + 1, column=index + 1).comment = comment


def show_comment(worksheet, contracts, brigades):
    for contract, comments in all_comments.items():
        for brigade, comment in comments.items():
            comment_index = [date_index for date_index, value in comment.items() if not isinstance(value, dict)]
            for day in comment_index:
                day_comment = Comment(comment[day].values[0], "")
                for index, cell in enumerate(worksheet[1]):
                    if int(day) == cell.value:
                        worksheet.cell(row=contracts.index(contract)+1, column=index+1).comment = day_comment


def show_mistakes_xlxs():
    file_name = 'output.xlsx'
    workbook = load_workbook(file_name)
    worksheet = workbook.active

    google_sheet_values = google_table_read()

    contracts = [cell.value for cell in worksheet["A"]]
    materials = [cell.value for cell in worksheet["C"]]
    brigades = [cell.value for cell in worksheet["B"]]

    for google_index, google_values in google_sheet_values.iterrows():
        contracts_sheet_indices = [sheet_index for sheet_index, sheet_contract_value in enumerate(contracts) if sheet_contract_value == google_values[0]]
        # for contract_sheet_index in contracts_sheet_indices:
        for material_sheet_index in contracts_sheet_indices:
            if google_values[0] == contracts[material_sheet_index]:
                google_material_name = google_values[1]
                brigade = brigades[material_sheet_index]
                if google_material_name == "ХП-Спрей" and materials[material_sheet_index] == " ХП-Спрей":
                    material_inaccuracy_calc(worksheet, material_sheet_index, google_values)
                elif google_material_name == "ТП-Спрей" and materials[material_sheet_index] == " ТП-Спрей":
                    material_inaccuracy_calc(worksheet, material_sheet_index, google_values)

                try:
                    if materials[material_sheet_index].split(' ')[1] == google_material_name.split(' ')[0] and (brigade.startswith('П') or brigade.startswith('К') or brigade.startswith('У')) and google_material_name.split(' ')[1] == 'лин':
                        material_inaccuracy_calc(worksheet, material_sheet_index, google_values)
                    elif materials[material_sheet_index].split(' ')[1] == google_material_name.split(' ')[0] and brigade.startswith('Р') and google_material_name.split(' ')[1] == 'руч':
                        material_inaccuracy_calc(worksheet, material_sheet_index, google_values)
                except IndexError:
                    pass

                if google_material_name == "СШк" and materials[material_sheet_index] == " СШ крупные":
                    material_inaccuracy_calc(worksheet, material_sheet_index, google_values)
                elif google_material_name == "СШм" and materials[material_sheet_index] == " СШ мелкие":
                    material_inaccuracy_calc(worksheet, material_sheet_index, google_values)
    show_comment(worksheet, contracts, brigades)
    for i in range(1, worksheet.max_column + 1):
        letter = get_column_letter(i)
        if letter != "A" and letter != "B" and letter != "C":
            worksheet.column_dimensions[letter].width = 8
        elif letter == "A":
            worksheet.column_dimensions[letter].width = 30
        elif letter == "B":
            worksheet.column_dimensions[letter].width = 20
        elif letter == "C":
            worksheet.column_dimensions[letter].width = 15
    workbook.save('output.xlsx')


def create_xlxs(data):
    # rows = []
    # date = [i for i in range(1, 32)]
    # for key, value in data.items():
    #     contract = key
    #     for person, details in value.items():
    #         row = [contract, person]
    #         rows.append(row)
    #
    # # Создание DataFrame
    # df = pd.DataFrame(rows, columns=['Contract', 'Person'])
    #
    # # Экспорт DataFrame в Excel
    # df.to_excel('output.xlsx', index=False)
    rows = []
    date = [i for i in range(1, 32)]
    rows.append(['Контракт', 'Бригада', 'Материал'] + date)
    for contract, contract_info in data.items():
        for brigade, details in contract_info.items():
            for day, material in details.items():
                for material_name, material_count in material.items():
                    short_material_name = material_name.split('/')[3]
                    row = [contract, brigade, short_material_name]
                    #   row.extend(("") * (int(day) - 1))
                    #   row_equal = any(set(row).issubset(sublist) for sublist in rows)
                    row_equal_index = next((i for i, sublist in enumerate(rows) if sublist[:3] == row), False)
                    if row_equal_index:
                        values = [None for _ in range(int(day) - 1)]
                        rows[row_equal_index].extend(values)
                        rows[row_equal_index].insert((int(day) + 2), details.get(day).get(material_name))
                    else:
                        values = [None for _ in range(int(day) - 1)]
                        row.extend(values)
                        row.append(details.get(day).get(material_name))  # Получаем данные для каждого дня
                        rows.append(row)

    df = pd.DataFrame(rows)

    file_name = 'output.xlsx'
    df.to_excel(file_name, index=False, header=False)
    show_mistakes_xlxs()


def out_material_calculatiom(work_material, out_material, brigades_materials_used_info, brigade_day_info, data, contract,
                             date, brigade, brigade_all_category_material_count):
    brigade_one_category_material_count = \
    brigade_day_info[data["Material stock picking out items/Категория материала/Отображаемое Имя"] == out_material][
        data["Material stock picking out items/Контракт/Отображаемое Имя"] == contract][
        "Material stock picking out items/Количество в базовых"].sum()
    brigade_material_work_square = \
    brigade_day_info[data["Event Prod Item/Категория материала/Отображаемое Имя"] == work_material][
        data["Event Prod Item/Контракт/Отображаемое Имя"] == contract]["Event Prod Item/Площадь, м²"].sum()
    try:
        if work_material in brigades_materials_used_info[contract][brigade][date]:
            brigade_all_category_material_count += brigade_one_category_material_count
        else:
            brigade_all_category_material_count = brigade_one_category_material_count
    except KeyError as e:
        brigades_materials_used_info[contract][brigade][date] = {}
        brigade_all_category_material_count = brigade_one_category_material_count

    return brigade_all_category_material_count, brigade_material_work_square


def csv_edit():
    file_name = "square&material.csv"
    data = pd.read_csv(file_name)
    # data.dropna(subset=["Event Prod Item/Контракт/Отображаемое Имя", 'Material stock picking out items/Контракт/Отображаемое Имя'], inplace=True)
    mask = data['Event Prod Item/Контракт/Отображаемое Имя'].isna() & data['Material stock picking out items/Контракт/Отображаемое Имя'].isna()
    data = data[~mask]
    data.drop(columns=["External ID"], inplace=True)
    data['Бригада/Отображаемое Имя'].fillna(method='ffill', inplace=True)
    data['Начало смены'] = pd.to_datetime(data['Начало смены'])
    data['Начало смены'].fillna(method='ffill', inplace=True)
    data['Начало смены'] = data['Начало смены'].dt.strftime('%e').str.strip()
    # data['Активности/Контракты/Субрегион'].fillna(method='ffill', inplace=True)
    data['Event Prod Item/Контракт/Отображаемое Имя'].fillna(method='ffill', inplace=True)
    contracts = data['Event Prod Item/Контракт/Отображаемое Имя'].drop_duplicates().dropna()

    comments = {}
    brigades_materials_used_info = {}

    # for brigade in brigades:
    #     brigades_materials_used_info[brigade] = {}
    #     for date in dates:
    #         brigade_day_info = data[data["Бригада/Отображаемое Имя"] == brigade][data["Начало смены"] == date]
    #         category_work_materials = brigade_day_info["Event Prod Item/Категория материала/Отображаемое Имя"].drop_duplicates()
    #         category_out_materials = brigade_day_info["Material stock picking out items/Категория материала/Отображаемое Имя"].drop_duplicates()
    #         contracts = brigade_day_info["Event Prod Item/Контракт/Отображаемое Имя"].drop_duplicates().dropna()
    #         brigades_materials_used_info[brigade][date] = {}
    #         brigade_all_category_material_count = 0
    #         for work_material in category_work_materials:
    #             for out_material in category_out_materials:
    #                     try:
    #                         if work_material in out_material or out_material == "Все / Материалы / Добавка / СШ крупные" or out_material == "Все / Материалы / Добавка / СШ мелкие":
    #                             for contract in contracts:
    #                                 brigade_one_category_material_count = brigade_day_info[data["Material stock picking out items/Категория материала/Отображаемое Имя"] == out_material][data["Material stock picking out items/Контракт/Отображаемое Имя"] == contract]["Material stock picking out items/Количество в базовых"].sum()
    #                                 brigade_material_work_square = brigade_day_info[data["Event Prod Item/Категория материала/Отображаемое Имя"] == work_material][data["Event Prod Item/Контракт/Отображаемое Имя"] == contract]["Event Prod Item/Площадь, м²"].sum()
    #                                 try:
    #                                     if work_material in brigades_materials_used_info[brigade][date][contract]:
    #                                         brigade_all_category_material_count += brigade_one_category_material_count
    #                                     else:
    #                                         brigade_all_category_material_count = brigade_one_category_material_count
    #                                 except KeyError as e:
    #                                     brigades_materials_used_info[brigade][date][contract] = {}
    #                                     brigade_all_category_material_count = brigade_one_category_material_count
    #                                 if out_material == "Все / Материалы / Добавка / СШ крупные" or out_material == "Все / Материалы / Добавка / СШ мелкие":
    #                                     brigade_flow_rate_metr = brigade_one_category_material_count / brigade_material_work_square
    #                                     brigades_materials_used_info[brigade][date][contract][out_material] = brigade_flow_rate_metr.round(2)
    #                                 else:
    #                                     brigade_flow_rate_metr = brigade_all_category_material_count / brigade_material_work_square
    #                                     brigades_materials_used_info[brigade][date][contract][work_material] = brigade_flow_rate_metr.round(2)
    #                     except TypeError as e:
    #                         continue
    for contract in contracts:
        dates = data[data["Event Prod Item/Контракт/Отображаемое Имя"] == contract]['Начало смены'].drop_duplicates()
        brigades_materials_used_info[contract] = {}
        comments[contract] = {}
        brigades = data[data["Event Prod Item/Контракт/Отображаемое Имя"] == contract]['Бригада/Отображаемое Имя'].drop_duplicates()
        for brigade in brigades:
            brigades_materials_used_info[contract][brigade] = {}
            comments[contract][brigade] = {}
            for date in dates:
                brigade_day_info = data[data["Бригада/Отображаемое Имя"] == brigade][data["Начало смены"] == date]
                category_work_materials = brigade_day_info[data["Event Prod Item/Контракт/Отображаемое Имя"] == contract]["Event Prod Item/Категория материала/Отображаемое Имя"].drop_duplicates().dropna()
                category_out_materials = brigade_day_info[data["Material stock picking out items/Контракт/Отображаемое Имя"] == contract]["Material stock picking out items/Категория материала/Отображаемое Имя"].drop_duplicates()
                brigades_materials_used_info[contract][brigade][date] = {}
                comments[contract][brigade][date] = {}
                brigade_all_category_material_count = 0
                for work_material in category_work_materials:
                    for out_material in category_out_materials:
                        try:
                            if work_material in out_material and "Спрей" not in out_material:
                                    brigade_all_category_material_count, brigade_material_work_square = out_material_calculatiom(work_material, out_material, brigades_materials_used_info, brigade_day_info, data, contract, date, brigade, brigade_all_category_material_count)
                                    brigade_flow_rate_metr = brigade_all_category_material_count / brigade_material_work_square
                                    brigades_materials_used_info[contract][brigade][date][work_material] = brigade_flow_rate_metr.round(2)

                            if work_material in out_material and "Спрей" in out_material and "Спрей" in work_material:
                                    brigade_all_category_material_count, brigade_material_work_square = out_material_calculatiom(work_material, out_material, brigades_materials_used_info, brigade_day_info, data, contract, date, brigade, brigade_all_category_material_count)
                                    brigade_flow_rate_metr = brigade_all_category_material_count / brigade_material_work_square
                                    brigades_materials_used_info[contract][brigade][date][work_material] = brigade_flow_rate_metr.round(2)

                            if out_material == "Все / Материалы / Добавка / СШ мелкие":
                                brigade_one_category_material_count = brigade_day_info[data["Material stock picking out items/Категория материала/Отображаемое Имя"] == out_material][data["Material stock picking out items/Контракт/Отображаемое Имя"] == contract][
                                    "Material stock picking out items/Количество в базовых"].sum()
                                mask = data["Event Prod Item/Категория материала/Отображаемое Имя"].notna() & data["Event Prod Item/Категория материала/Отображаемое Имя"].str.contains("Краска", na=False)
                                brigade_material_work_square = brigade_day_info[mask][data["Event Prod Item/Контракт/Отображаемое Имя"] == contract]["Event Prod Item/Площадь, м²"].sum()
                                brigade_flow_rate_metr = brigade_one_category_material_count / brigade_material_work_square
                                brigades_materials_used_info[contract][brigade][date][out_material] = brigade_flow_rate_metr.round(2)
                            if out_material == "Все / Материалы / Добавка / СШ крупные":
                                brigade_one_category_material_count = brigade_day_info[data["Material stock picking out items/Категория материала/Отображаемое Имя"] == out_material][data["Material stock picking out items/Контракт/Отображаемое Имя"] == contract][
                                    "Material stock picking out items/Количество в базовых"].sum()
                                mask = data["Event Prod Item/Категория материала/Отображаемое Имя"].notna() & data["Event Prod Item/Категория материала/Отображаемое Имя"].str.contains("ТП|ХП", na=False)
                                brigade_material_work_square = brigade_day_info[mask][data["Event Prod Item/Контракт/Отображаемое Имя"] == contract]["Event Prod Item/Площадь, м²"].sum()
                                brigade_flow_rate_metr = brigade_one_category_material_count / brigade_material_work_square
                                brigades_materials_used_info[contract][brigade][date][out_material] = brigade_flow_rate_metr.round(2)
                            comments[contract][brigade][date] = brigade_day_info.apply(lambda row: ' '.join([str(x) for x in [row['Активности/Доп. информация'],
                                    row['Сообщения с сайта/Содержание']] if x is not np.nan]) if any(x is not np.nan for x in [row['Активности/Доп. информация'], row['Сообщения с сайта/Содержание']])
                                    else np.nan, axis=1)
                            if comments[contract][brigade][date].notna().any():
                                comments[contract][brigade][date] = comments[contract][brigade][date].dropna()
                            else:
                                comments[contract][brigade][date] = {}
                        except TypeError as e:
                            continue
    global all_comments
    all_comments.update(comments)
    # pprint(brigades_materials_used_info)
    create_xlxs(brigades_materials_used_info)


csv_edit()
