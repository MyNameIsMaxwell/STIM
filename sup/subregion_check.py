import pandas as pd
from openpyxl import Workbook, load_workbook


def subregion_show():
    file_name = 'economic_report.csv'

    data = pd.read_csv(file_name)
    subregion_info = data.loc[:, ['Активности/Контракты/Субрегион', 'Бригада/Отображаемое Имя', 'Начало смены']].dropna(subset=['Активности/Контракты/Субрегион'])
    subregion_info['Начало смены'] = pd.to_datetime(data['Начало смены'])
    subregion_info['Начало смены'] = subregion_info['Начало смены'].dt.strftime('%e')
    subregion_info = subregion_info.drop_duplicates()
    brigades = subregion_info.drop_duplicates(subset=['Бригада/Отображаемое Имя'])['Бригада/Отображаемое Имя']
    wb = load_workbook('subregion.xlsx')
    ws = wb.active

    for brigade in brigades:
        for row in ws.iter_cols(min_col=0, max_col=1):
            for cell in row:
                if cell.value == brigade:
                    row_number = cell.row
                    for day in subregion_info[(subregion_info['Бригада/Отображаемое Имя'] == brigade)]['Начало смены']:
                        for row in ws.iter_rows(min_row=0, max_row=1):
                            for cell in row:
                                if cell.value == int(day):
                                    column_number = cell.column
                                    value = subregion_info[(subregion_info['Бригада/Отображаемое Имя'] == brigade) &
                                                   (subregion_info['Начало смены'] == day)]['Активности/Контракты/Субрегион']
                                    if not value.empty:
                                        ws.cell(row=row_number,
                                                column=column_number,
                                                value='')
                                        ws.cell(row=row_number,
                                                column=column_number,
                                                value=value.to_string(index=False))
                                    break
    wb.save('subregion.xlsx')


subregion_show()


