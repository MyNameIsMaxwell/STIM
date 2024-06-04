import time
from pprint import pprint

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def fuel_for_one_car():
    # Путь к CSV-файлу
    # csv_file = input('Введите имя_файла.csv: ')
    csv_file = 'roadmark_mrp_shift.work_shift.csv'

    # Создание нового Excel-файла и выбор активного листа
    workbook = Workbook()
    worksheet = workbook.active

    # Чтение данных из CSV-файла с помощью pandas
    data = pd.read_csv(csv_file)

    filtered_data = data[
        (data['Расход топлива/Оборудование/Наименование оборудования'] == 'Д01 MAN АН 8499-1 (демаркиратор)') |
        (data['Начало смены'].notnull())
        ]

    # Запись данных в Excel-таблицу
    for index, row in filtered_data.iterrows():
        worksheet.append(row.tolist())

    # Сохранение Excel-файла
    excel_file = 'новый_файл.xlsx'
    workbook.save(excel_file)


def fuel_for_all():
    file_name = 'roadmark_mrp_shift.work_shift.xls'
    new_file_name = 'ГСМ odoo сентябрь(обр).xlsx'

    df = pd.read_excel(file_name)

    # Заполните пустые ячейки в первом столбце данными из предыдущей строки
    df.iloc[:, 1].fillna(method='ffill', inplace=True)
    df.iloc[:, 2].fillna(method='ffill', inplace=True)
    df.iloc[:, 3].fillna(method='ffill', inplace=True)

    df.to_excel(new_file_name, index=False)

    time.sleep(1)

    workbook = load_workbook(new_file_name)
    worksheet = workbook.active

    # sort_conditions = [
    #     {'column': 'B', 'ascending': True},  # Сортировка по столбцу A (по возрастанию)
    #     {'column': 'E', 'ascending': True},  # Сортировка по столбцу B (по убыванию)
    #     {'column': 'G', 'ascending': True}  # Сортировка по столбцу C (по возрастанию)
    # ]
    #
    # for condition in sort_conditions:
    #     # Выберите диапазон ячеек для сортировки (например, столбец A и строки с 2 по последнюю)
    #     range_to_sort = f'{condition["column"]}2:{condition["column"]}' + str(worksheet.max_row)
    #
    #     # Отсортируйте ячейки в указанном диапазоне
    #     worksheet.auto_filter.ref = range_to_sort
    #     worksheet.auto_filter.add_sort_condition(f'{condition["column"]}2:{condition["column"]}',
    #                                       descending=not condition["ascending"])

    # worksheet.row_dimensions.group(worksheet.min_row, worksheet.max_row, outline_level=3)

    for i in range(1, worksheet.max_column + 1):
        letter = get_column_letter(i)

        worksheet.column_dimensions[letter].width = 30

    workbook.save(new_file_name)


def second_time_filling():
    file_name = 'fuel_april.csv'

    data = pd.read_csv(file_name)

    # for i in range(1, worksheet.max_column + 1):
    #     letter = get_column_letter(i)
    #
    #     worksheet.column_dimensions[letter].width = 30
    filtered_data = data[
        (data['Расход топлива/Оборудование/Наименование оборудования'] != '')
    ]
    fuel_info = dict()
    for day_data in filtered_data.iterrows():
        for brigade_info in day_data:
            if not isinstance(brigade_info, int):
                # car = brigade_info["Расход топлива/Оборудование/Наименование оборудования"]
                if not is_nan(brigade_info["Бригада"]) and brigade_info["Бригада"] not in fuel_info and not "Суб" in \
                                                                                                            brigade_info[
                                                                                                                "Бригада"]:
                    fuel_info[brigade_info["Бригада"]] = dict()
                if brigade_info["Бригада"] in fuel_info:
                    brigade = brigade_info["Бригада"]
                if brigade_info["Расход топлива/Заправка / Расход"] == "Заправка" and not fuel_info.get(brigade).get(
                        f'{brigade_info["Расход топлива/Оборудование/Наименование оборудования"]}-Заправка'):
                    fuel_info[brigade][
                        f'{brigade_info["Расход топлива/Оборудование/Наименование оборудования"]}-Заправка'] = float(
                        brigade_info["Расход топлива/Количество"])

                elif brigade_info["Расход топлива/Заправка / Расход"] == "Заправка" and fuel_info.get(brigade).get(
                        f'{brigade_info["Расход топлива/Оборудование/Наименование оборудования"]}-Заправка'):
                    fuel_info[brigade][
                        f'{brigade_info["Расход топлива/Оборудование/Наименование оборудования"]}-Заправка'] += float(
                        brigade_info["Расход топлива/Количество"])

                elif brigade_info["Расход топлива/Заправка / Расход"] == "Расход" and not fuel_info.get(brigade).get(
                        f'{brigade_info["Расход топлива/Оборудование/Наименование оборудования"]}-Расход'):
                    fuel_info[brigade][
                        f'{brigade_info["Расход топлива/Оборудование/Наименование оборудования"]}-Расход'] = float(
                        brigade_info["Расход топлива/Количество"])

                elif brigade_info["Расход топлива/Заправка / Расход"] == "Расход" and fuel_info.get(brigade).get(
                        f'{brigade_info["Расход топлива/Оборудование/Наименование оборудования"]}-Расход'):
                    fuel_info[brigade][
                        f'{brigade_info["Расход топлива/Оборудование/Наименование оборудования"]}-Расход'] += float(
                        brigade_info["Расход топлива/Количество"])

                try:
                    if not fuel_info.get(brigade_info["Бригада"]).get("Регион") and not isinstance(
                            fuel_info.get(brigade_info["Бригада"]).get("Регион"), set):
                        fuel_info[brigade_info["Бригада"]]["Регион"] = set()
                    elif isinstance(fuel_info.get(brigade_info["Бригада"]).get("Регион"), set):
                        fuel_info[brigade_info["Бригада"]]["Регион"].add(brigade_info["Регион"])
                except AttributeError:
                    """no info"""
                    pass

    # pprint(fuel_info)

    records = []

    # for name, values in fuel_info.items():
    #     region = values.get('Регион', '')
    #     machines = [key for key in values if key != 'Регион']
    #     if not records or name not in records[-1]:
    #         records.append((name, region))
    #     for machine in machines:
    #         value = values[machine]
    #         records.append(('', '', machine, value))

    for name, values in fuel_info.items():
        region = values.get('Регион', '')
        for machine, value in values.items():
            if machine != 'Регион':
                records.append((name, region, machine, value))

    df = pd.DataFrame(records, columns=['Бригада', 'Регион', 'Машина', 'Значение'])

    # df = pd.DataFrame.from_dict(fuel_info, orient='index')
    # df = df.stack(level=0).reset_index()
    # df.columns = ['Бригада', 'Регион', 'Машина', 'Значение']

    df.to_excel('Топливо.xlsx', index=False)
    time.sleep(1)

    wb = load_workbook('Топливо.xlsx')
    ws = wb.active

    ws.column_dimensions['A'].width = 21.5
    ws.column_dimensions['B'].width = 19
    ws.column_dimensions['C'].width = 37.7

    wb.save('Топливо.xlsx')


def is_nan(x):
    return (x != x)


# fuel_for_all()

second_time_filling()
