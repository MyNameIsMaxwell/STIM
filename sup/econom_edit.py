import json
import os
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment

import pandas as pd


def column_size(sheet):
    for i in range(2, 100):
        letter = get_column_letter(i)
        sheet.column_dimensions[letter].width = 15
    sheet.column_dimensions['A'].width = 35


def column_alignment(sheet):
    # выравнивание значений по центру
    for cell in range(1, sheet.max_column+1):
        cell_info = sheet.cell(row=1, column=cell)
        if cell_info.value:
            cell_info.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
def create_table():
    with open(os.path.join("D:\PyCharm_Projects\sup\sup\Экономисты", 'material_info.json'), 'r', encoding='utf-8') as file:
        data = json.load(file)

    # Создание нового файла Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    column_size(sheet)

    row = 2
    sheet.cell(row=1, column=2).value = "Контракт"
    sheet.cell(row=1, column=3).value = "Страна"
    sheet.cell(row=1, column=4).value = "Регион"
    sheet.cell(row=1, column=5).value = "Общий регион"
    sheet.cell(row=1, column=6).value = "Вид материала"
    sheet.cell(row=1, column=7).value = "Основной/Гарантия"
    sheet.cell(row=1, column=8).value = "Месяц"
    sheet.cell(row=1, column=9).value = "Работы"
    sheet.cell(row=1, column=10).value = "Вид работ"
    sheet.cell(row=1, column=11).value = "Тип бригады"
    sheet.cell(row=1, column=12).value = "Бригада"
    sheet.cell(row=1, column=12).value = "ЗП"
    sheet.cell(row=1, column=13).value = "Выручка"
    sheet.cell(row=1, column=14).value = "Количество ХП/ТП/АК, кг"
    sheet.cell(row=1, column=15).value = "Стоимость ХП/ТП/АК"
    sheet.cell(row=1, column=16).value = "Количество разбавителя"
    sheet.cell(row=1, column=17).value = "Стоимость разбавителя"
    sheet.cell(row=1, column=18).value = "Количество СШ больших"
    sheet.cell(row=1, column=19).value = "Стоимость СШ больших"
    sheet.cell(row=1, column=20).value = "Количество СШ мелких"
    sheet.cell(row=1, column=21).value = "Стоимость СШ мелких"

    column_alignment(sheet)


    # for contract, values in data.items():
    #     sheet.cell(row=row, column=1).value = contract
    #     # sheet.cell(row=row, column=contract_column-1).value = values['contract_material']
    #     for brigade, materials in values['brigades_material'].items():
    #         for material, info in materials.items():
    #             sheet.cell(row=row+1, column=2)
    #             sheet.cell(row=row+1, column=2).value = contract
    #             sheet.cell(row=row, column=9).value = material
    #             sheet.cell(row=row+1, column=1).value = brigade
    #
    #             row += 3

    # Сохранение Excel-файла
    workbook.save(os.path.join("D:\PyCharm_Projects\sup\sup\Экономисты", 'output.xlsx'))


def material_search():
    csv_mat = 'econom_mat.csv'
    material_file_path = os.path.join("D:\PyCharm_Projects\sup\sup\Экономисты", csv_mat)

    df = pd.read_csv(material_file_path)

    brigade_list = df['Бригада/Отображаемое Имя'].drop_duplicates().values.tolist()
    contracts_list = df['Контракт/Отображаемое Имя'].drop_duplicates().values.tolist()
    materials_list = df['Материал/Отображаемое Имя'].drop_duplicates().values.tolist()

    ak_standart = {'Все / Материалы / Основной / Краска / Белый / Стандарт'}
    ak_sprinter = {'Все / Материалы / Основной / Краска / Белый / Спринтер'}
    ak_colorful = {'Все / Материалы / Основной / Краска / Желтый', 'Все / Материалы / Основной / Краска / Желтый / Стандарт', 'Все / Материалы / Основной / Краска / Цветной / Зеленый', 'Все / Материалы / Основной / Краска / Цветной / Красный', 'Все / Материалы / Основной / Краска / Цветной / Оранжевый', 'Все / Материалы / Основной / Краска / Цветной / Черный'}
    HP_white = {'Все / Материалы / Основной / ХП / Белый'}
    HP_colorful = {'Все / Материалы / Основной / ХП / Желтый', 'Все / Материалы / Основной / ХП / Цветной / Зеленый', 'Все / Материалы / Основной / ХП / Цветной / Красный', 'Все / Материалы / Основной / ХП / Цветной / Черный'}
    TP_white = {'Все / Материалы / Основной / ТП / Белый', 'Все / Материалы / Основной / ТП / Белый / В6', 'Все / Материалы / Основной / ТП / Белый / В7'}
    TP_colorful = {'Все / Материалы / Основной / ТП / Желтый'}
    diluent = {'Все / Материалы / Добавка / Разбавитель'}
    glassballs_big = {'Все / Материалы / Добавка / СШ крупные'}
    glassballs_small = {'Все / Материалы / Добавка / СШ мелкие'}

    all_materials = {
        "Краска": [ak_standart, ak_sprinter, ak_colorful],
        "ХП": [HP_white, HP_colorful],
        "ТП": [TP_white, TP_colorful],
        "Растворитель": [diluent],
        "Стеклошарики крупные": [glassballs_big],
        "Стеклошарики мелкие": [glassballs_small],
        }

    white_colorful_category = {
          "Белый": ['Все / Материалы / Основной / ХП / Белый', 'Все / Материалы / Основной / ТП / Белый', 'Все / Материалы / Основной / ТП / Белый / В6', 'Все / Материалы / Основной / ТП / Белый / В7'],
          "Цветной": ['Все / Материалы / Основной / Краска / Желтый', 'Все / Материалы / Основной / Краска / Желтый / Стандарт', 'Все / Материалы / Основной / Краска / Цветной / Зеленый', 'Все / Материалы / Основной / Краска / Цветной / Красный', 'Все / Материалы / Основной / Краска / Цветной / Оранжевый', 'Все / Материалы / Основной / Краска / Цветной / Черный', 'Все / Материалы / Основной / ХП / Желтый', 'Все / Материалы / Основной / ХП / Цветной / Зеленый', 'Все / Материалы / Основной / ХП / Цветной / Красный', 'Все / Материалы / Основной / ХП / Цветной / Черный', 'Все / Материалы / Основной / ТП / Желтый'],
          "Стандарт": ['Все / Материалы / Основной / Краска / Белый / Стандарт'],
          "Спринтер": ['Все / Материалы / Основной / Краска / Белый / Спринтер'],
      }

    full_material_contracts = {}

    for contract in contracts_list:
        contract_materials_count = {}
        brigade_materials_count = {}
        contract_info_filter = (df['Контракт/Отображаемое Имя'] == contract)
        brigades_on_contract = df[contract_info_filter]['Бригада/Отображаемое Имя'].drop_duplicates().values.tolist()
        for brigade in brigades_on_contract:
            brigade_on_contract_info_filter = (contract_info_filter & (df['Бригада/Отображаемое Имя'] == brigade))
            materials_on_contract = df[brigade_on_contract_info_filter]['Категория материала/Отображаемое Имя'].drop_duplicates().values.tolist()
            brigade_materials_count[brigade] = {}
            for material in materials_on_contract:
                full_material_contracts[contract] = {}  # создание контракта и инфы о нем
                materials_brigade_count_filter = (brigade_on_contract_info_filter & (df['Категория материала/Отображаемое Имя'] == material))
                materials_contract_count_filter = (contract_info_filter & (df['Категория материала/Отображаемое Имя'] == material))
                material_name = next((name for name, material_full_name in all_materials.items() for mat in material_full_name if material in mat), None)
                material_contract_count = df[materials_contract_count_filter]['Количество в базовых'].sum()
                material_brigade_count = df[materials_brigade_count_filter]['Количество в базовых'].sum()
                white_or_colorful = next((name for name, material_full_name in white_colorful_category.items() for mat in material_full_name if material in mat), None)
                contract_materials_count[material_name] = material_contract_count
                brigade_materials_count[brigade][material_name] = [material_brigade_count, white_or_colorful]
            full_material_contracts[contract]["brigades_material"] = brigade_materials_count
        full_material_contracts[contract]["contract_material"] = contract_materials_count

    json_file_name = "material_info.json"
    folder_name = "D:\PyCharm_Projects\sup\sup\Экономисты"
    json_file_path = os.path.join("D:\PyCharm_Projects\sup\sup\Экономисты", json_file_name)

    if not os.path.exists(folder_name):
        os.makedirs(folder_name)

    with open(json_file_path, "w", encoding='utf-8') as json_file_name:
        json.dump(full_material_contracts, json_file_name, indent=2, ensure_ascii=False)


def square_count():
    csv_square = 'econom_square.csv'
    material_file_path = os.path.join("D:\PyCharm_Projects\sup\sup\Экономисты", csv_square)

    df = pd.read_csv(material_file_path)

    contracts_list = df['Контракт/Отображаемое Имя'].drop_duplicates().values.tolist()

    full_square_contracts = {}

    for contract in contracts_list:
        contract_square_count = {}
        brigade_square_count = {}
        contract_info_filter = (df['Контракт/Отображаемое Имя'] == contract)
        brigades_on_contract = df[contract_info_filter]['Бригада/Отображаемое Имя'].drop_duplicates().values.tolist()
        for brigade in brigades_on_contract:
            brigade_on_contract_info_filter = (contract_info_filter & (df['Бригада/Отображаемое Имя'] == brigade))
            materials_on_contract = df[brigade_on_contract_info_filter]['Категория материала/Отображаемое Имя'].drop_duplicates().values.tolist()
            work_type = df[brigade_on_contract_info_filter]['Вид работ'].drop_duplicates().values.tolist()
            brigade_square_count[brigade] = {}
            for material in materials_on_contract:
                contract_square_count[material] = {}
                brigade_square_count[brigade][material] = {}
                for type in work_type:
                    full_square_contracts[contract] = {}  # создание контракта и инфы о нем
                    materials_brigade_count_filter = (brigade_on_contract_info_filter & (df['Категория материала/Отображаемое Имя'] == material) & (df['Вид работ'] == type))
                    materials_contract_count_filter = (contract_info_filter & (df['Категория материала/Отображаемое Имя'] == material) & (df['Вид работ'] == type))
                    # material_name = next((name for name, material_full_name in all_materials.items() for mat in material_full_name if material in mat), None)
                    square_contract_count = df[materials_contract_count_filter]['Площадь, м²'].sum()
                    square_brigade_count = df[materials_brigade_count_filter]['Площадь, м²'].sum()
                    contract_square_count[material][type] = square_contract_count
                    brigade_square_count[brigade][material][type] = [square_brigade_count, f'{material}-{type[:3]}']
            full_square_contracts[contract]["brigades_material"] = brigade_square_count
        full_square_contracts[contract]["contract_material"] = contract_square_count
        full_square_contracts[contract]["contract_material"]["total"] = df[contract_info_filter]['Площадь, м²'].sum()

        json_file_name = "square_info.json"
        folder_name = "D:\PyCharm_Projects\sup\sup\Экономисты"
        json_file_path = os.path.join("D:\PyCharm_Projects\sup\sup\Экономисты", json_file_name)

        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

        with open(json_file_path, "w", encoding='utf-8') as json_file_name:
            json.dump(full_square_contracts, json_file_name, indent=2, ensure_ascii=False)


def region():
    pass

def date():
    pass


if __name__ == "__main__":
    material_search()
    # create_table()
    # square_count()